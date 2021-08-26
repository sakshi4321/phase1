#!/usr/bin/env python
import os
import shutil
import csv
import openpyxl
import threading
from openpyxl.reader.excel import load_workbook
from flask import Flask, render_template, request, \
    Response, send_file, redirect, url_for,flash, make_response 
from camera import Camera

from flask import send_file, send_from_directory, safe_join, abort,session
from flask_sqlalchemy import SQLAlchemy
from flask import Blueprint,flash
from flask_login import login_user, logout_user, login_required, current_user,LoginManager, login_manager
from flask_login import UserMixin
import sqlalchemy as db
import pandas as pd
#from datetime import datetime
from sqlalchemy import Column, Integer, DateTime
from sqlalchemy.ext.declarative import declarative_base


import pickle
import cv2
import random
import datetime
from datetime import datetime, date
import xlwt
import xlrd
from facenet_pytorch import MTCNN
from xlwt import Workbook 
from matplotlib import pyplot
#from mtcnn.mtcnn import MTCNN
from matplotlib.patches import Rectangle
from keras.models import load_model
from matplotlib.patches import Circle
import cv2
import numpy as np
from sklearn.preprocessing import Normalizer
from scipy.spatial.distance import cosine
import os
from werkzeug.security import generate_password_hash, check_password_hash

from xlutils.copy import copy
from sqlalchemy import and_, or_, not_
# results=db.session.query(Students,Course,Classes).\
# ... select_from(Students).join(Course).join(Classes).all()

from xlrd import open_workbook
#mysql://root:''@localhost/attendance
app = Flask(__name__)
#ap = Blueprint('main', __name__)
login = LoginManager(app)
app.config["SECRET_KEY"]="abc"
app.config["SQLALCHEMY_DATABASE_URI"]="sqlite:///attendance.db"
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
camera = None
#login_manager = LoginManager(app)
marked_courses=[]
db=SQLAlchemy(app)
# app.secret_key = "abc"


# lecs=db.Table('lecs',
#                 db.Column('class_id',db.Integer,db.ForeignKey('classes.class_id')),
#                 db.Column('lecture_id',db.Integer,db.ForeignKey('lectures.lecture_id')))


######################################### DATABASE ##################################################################
class Students(db.Model):
    roll_no=db.Column(db.String(200),primary_key=True)
    rank=db.Column(db.String(200),nullable=False)
    first_name=db.Column(db.String(200),nullable=False)
    last_name=db.Column(db.String(200),nullable=False)
    phone=db.Column(db.Integer(),nullable=False)
    attendance_id=db.relationship('Attendance_sys',backref='stud_attend')
    # room_id=db.Column(db.Integer,db.ForeignKey('classes.class_id'))
    course_sel=db.Column(db.Integer,db.ForeignKey('course.course_id'))
    id=db.relationship('arecord',backref='id',cascade = "all,delete, delete-orphan")

    


    # def __init__(self,roll_no,first_name,last_name,phone,room_id):
    #     self.roll_no=roll_no
    #     self.first_name=first_name
    #     self.last_name=last_name
    #     self.phone=phone
    #     self.room_id=room_id


# class Timetable(db.Model):
#     time_id=db.Column(db.Integer,primary_key=True)
#     start_time=db.Column(db.String(200),nullable=False)
#     end_time=db.Column(db.String(200),nullable=False)

class Course(db.Model):
    course_id=db.Column(db.Integer,primary_key=True)
    course_name=db.Column(db.String(200),unique=True,nullable=False)
    # stream=db.Column(db.String(200),nullable=False)
    courses=db.relationship('Students',backref='courses',cascade = "all,delete, delete-orphan")
    course_class=db.relationship('Classes',backref='course_class',cascade = "all,delete, delete-orphan")

class Classes(db.Model):
    class_id=db.Column(db.Integer,primary_key=True)
    classname=db.Column(db.String(200),unique=True,nullable=False)
    camera_name=db.Column(db.String(200),nullable=False)
    course_sel=db.Column(db.Integer,db.ForeignKey('course.course_id'))
    # roll=db.relationship('Students',backref='classroom')
    attendance_id=db.relationship('Attendance_sys',backref='cla_attend')
    # lectures=db.relationship('Lectures',secondary=lecs,backref=db.backref('subjects',lazy='dynamic'))

    # def __init__(self,class_id,classname):
    #     self.class_id=class_id
    #     self.classname=classname
    #     self.roll_no=roll_no
    #     self.lectures=lectures

# class timing(db.Model):
#     time_id=db.Column(db.Integer,primary_key=True)
#     time=db.Column(db.String(200))
    #minutes=db.Column(db.Integer)
class Lectures(db.Model):
    lecture_id=db.Column(db.Integer,primary_key=True)
    #Comment lecture_name and check how to show timings 
    lecture_name=db.Column(db.String(200),nullable=False)
    lecture_day=db.Column(db.String(200),nullable=False)
    lecture_time=db.Column(db.String(200),nullable=False)
    # lecture_datetime=db.Column(db.DateTime,default=datetime.utcnow())
    # lecture_start_time=db.Column(db.Time)
    # lecture_end_time=db.Column(db.Time)
    attendance_id=db.relationship('Attendance_sys',backref='lec_attend')

    # def __init__(self,lecture_id,class_id,lecture_date,lecture_start_time,lecture_end_time):
    #     self.lecture_id=lecture_id
    #     self.class_id=class_id
    #     self.lecture_date=lecture_date
    #     self.lecture_start_time=lecture_start_time
    #     self.lecture_end_time=lecture_end_time



# class Attendance_sys(db.Model):
#     attendance_id=db.Column(db.Integer,primary_key=True)
#     roll_no=db.Column(db.Integer,db.ForeignKey('students.roll_no'))
#     lecture_id=db.Column(db.Integer,db.ForeignKey('lectures.lecture_id'))
#     class_id=db.Column(db.Integer,db.ForeignKey('classes.class_id'))
#     present_absent=db.Column(db.Boolean, default=False, nullable=False)
# ####my
# class Attendance(db.Model):
#     course_id=db.Column(db.Integer,primary_key=True)
#     roll_no=db.Column(db.Integer,nullable=False)
#     lecture_id=db.Column(db.Integer,nullable=False)
#     class_id=db.Column(db.Integer,nullable=False)
#     present_absent=db.Column(db.Boolean, default=False, nullable=False)


    # def __init__(self,attendance_id,lecture_id,class_id,present_absent):
    #     self.attendance_id=attendance_id
    #     self.roll_no=roll_no
    #     self.lecture_id=lecture_id
    #     self.class_id=class_id
    #     self.present_absent=present_absent
"""engine = db.create_engine('sqlite:///test.sqlite') #Create test.sqlite automatically
connection = engine.connect()
metadata = db.MetaData()"""
#course='ic'
class arecord(db.Model):
    id_a=db.Column(db.String(225),db.ForeignKey('students.roll_no'))
    primkey=db.Column(db.Integer(),autoincrement=True, primary_key=True)
    name_a=db.Column(db.String(255), nullable=False)
    date=db.Column(db.Date, nullable=False, default=date.today())
    lecture_no=db.Column(db.Integer(), nullable=False)         
              
    attend=db.Column(db.Boolean(), default=False)
    
    """two=db.Column(db.Boolean(), default=False)
    three=db.Column(db.Boolean(), default=False)
    four=db.Column(db.Boolean(), default=False)
    five=db.Column(db.Boolean(), default=False)
    six=db.Column(db.Boolean(), default=False)
    seven=db.Column(db.Boolean(), default=False)
    eight=db.Column(db.Boolean(), default=False)"""
class Users(UserMixin,db.Model):
    id = db.Column(db.String(200), primary_key=True) # primary keys are required by SQLAlchemy
   
    
    password = db.Column(db.String(100))
    
db.create_all()

"""def __init__(self,id_a,name_a,attend):
        self.id_a=id_a
        self.name_a=name_a
        #self.date=date
        self.attend=attend"""
    #     self.present_absent=present_absent
    
"""              
emp = db.Table(course, 
              db.Column('Id', db.Integer(),primary_key=True),
              db.Column('name', db.String(255), nullable=False),
              db.Column('date', db.DateTime),
              
              
              db.Column('attendance', db.Boolean(), default=True)
              )"""
#metadata.create_all(engine) #Creates the table
########################################################################################################
###### Convert string to an Object
def str_to_class(str):
    return getattr(sys.modules[__name__], str)

    
def get_camera():
    global camera
    if not camera:
        camera = Camera()

    return camera
@login.user_loader
def load_user(id):
    return Users.query.get(id)
@app.route('/', methods=['GET', 'POST']) # define login page path
def login(): # define login page fucntion
    
    if request.method=='POST': 
        # if the request is POST the we check if the user exist and with te right password
        user = request.form.get("uname")
    
        password = request.form["psw"]
        # print(user)
        # print(password)
    
        dbuser = Users.query.filter_by(id=user).first()
        if dbuser is None:
            flash('Invalid Login')
            return redirect(url_for("login")) 
        #print(dbuser.)
        #if user==dbuser.id and password==dbuser.password: 
        else:
            if user==dbuser.id and password==dbuser.password: 
                #print("check")
                login_user(dbuser)
                return redirect(url_for('index'))
            else:
                flash('Invalid Login')
                return redirect(url_for("login"))

    
    return render_template('register.html')
    # if request.method=='POST': # if the request is a GET we return the login page
    #     # if the request is POST the we check if the user exist and with te right password
    #     user = request.form.get("uname")
    
    #     password = request.form["psw"]
    #     if user=="admin" and password=="admin@123":
    #         login_user(user, remember=remember)
    #         return redirect(url_for("index"))
    #     else :
    #         flash('Please check your login details and try again.')
    #         return redirect(url_for('login')) # if the user doesn't exist or password is wrong, reload the page
    #     # if the above check passes, then we know the user has the right credentials
        
    #     return redirect(url_for('login'))

@app.route('/logout') # define logout path
@login_required
def logout(): #define the logout function
    logout_user()
    return redirect(url_for('login'))

"""
@app.route('/')
def root():
    return redirect(url_for('index'))
"""
# @app.route('/', methods =["GET", "POST"])
# def image():
#     return render_template('register.html')

#    if request.method == "POST":
#        user = request.form.get("uname")
    
#        password = request.form["psw"]
#        if user=="admin" and password=="admin@123":
#             return redirect(url_for("index"))
#    session["a"]=first_name
#    session["c"]=last_name


#os.mkdir(str(first_name)+"_"+str(last_name))
#os.chdir(str(first_name)+"_"+str(last_name))
#return "Your name is "+first_name + last_name
"""with open('nameList.csv','w') as inFile:

writer = csv.DictWriter(inFile, fieldnames=fieldnames)


writer.writerow({'name': name, 'comment': comment})"""

       
    
######## Student Registration page
@app.route('/index/', methods =["GET", "POST"])
@login_required
def index():
    # students=Students.query.all()
    # classrooms=Classes.query.all()
    # for c in classrooms:
    #     obj=Students.query.filter_by(room_id=c.class_id).first()
    # students=db.session.query(Students,Classes).join(Classes).all()
    course=Course.query.all()
    students=db.session.query(Students,Course).join(Course).all()
    return render_template('index_new.html',students=students,course=course)

######### Courses Registration Page
@app.route('/courses/', methods =["GET", "POST"])
@login_required
def course_reg():
    courses=Course.query.all()
    return render_template('courses.html',courses=courses)

@app.route('/lectures/', methods =["GET", "POST"])
@login_required
def lecture_reg():
    lectures=Lectures.query.all()
    return render_template('lectures.html',lectures=lectures)

######### Class Registration page
@app.route('/classes/', methods =["GET", "POST"])
@login_required
def class_reg():
    classes=db.session.query(Classes,Course).join(Course).all()
    courses=Course.query.all()
    return render_template('classes.html',classes=classes,courses=courses)

########## Attendance Page
@app.route('/attendance', methods=['POST', 'GET'])
@login_required
def attendance_records():
    attendance = arecord.query.all()
    attendance2 = arecord.query.with_entities(arecord.name_a).distinct()
    attendance1 = arecord.query.with_entities(arecord.lecture_no).distinct()
    return render_template('attend_page.html',record=attendance,nRecord=attendance2,nRecord1=attendance1)

@app.route('/timings', methods=['POST', 'GET'])
def lecture_timings():
    time = timing.query.all()
    
    return render_template('lecture_timing.html',record=time)



######## Insert Student in the database
@app.route('/insert',methods=["GET","POST"])
def insert():
    if request.method=="POST":
        # roll_no=request.form["roll"]
        # first_name=request.form["first_name"]
        # last_name=request.form["last_name"]
        # phone=request.form["phone"]
        camera=get_camera()
        camera.stop_cam()
        roll_no=session.get("roll_no")
        rank=session.get("rank")
        first_name=session.get("first_name")
        last_name=session.get("last_name")
        phone=session.get("phone")
        course=session.get("course")
        # class_name=session.get("class_name")
        # class_name_1=Classes.query.filter_by(classname=class_name).first()

        # courses=Course(course_name=course)
        # db.session.add(courses)
        # db.session.commit()

        course_name=Course.query.filter_by(course_name=course).first()
        
        students=Students(roll_no=roll_no,rank=rank,first_name=first_name,last_name=last_name,phone=phone,courses=course_name)
        db.session.add(students)
        db.session.commit()
        flash("Student Added Sucessfully!!")
        return redirect(url_for("index"))


###### Insert course in the database
@app.route('/course_insert',methods=["POST","GET"])
def insert_course():
    if request.method=="POST":
        course_name=request.form["course_name"]

        stud_course=Course.query.all()

        for x in stud_course:
            if course_name==x.course_name:
                flash(1)
                return redirect(url_for("course_reg"))

        courses=Course(course_name=course_name)
        db.session.add(courses)
        db.session.commit()
        flash("Course Added Successfully!!")
        return redirect(url_for('course_reg'))
##########insert lecture timing
@app.route('/timing_insert',methods=["POST","GET"])
def insert_time():
    if request.method=="POST":
        h=request.form["hours"]
        

        stud_course=timing.query.all()

        for x in stud_course:
            if h==x.time :
                flash(1)
                return redirect(url_for("lecture_timings"))

        courses=timing(time=h)
        db.session.add(courses)
        db.session.commit()
        flash("Lecture Timing Added Successfully!!")
        return redirect(url_for('lecture_timings'))
###### Insert class in the database
@app.route('/class_insert',methods=["POST","GET"])
def insert_class():
    if request.method=="POST":
        class_name=request.form["class_name"]
        camera_name=request.form["camera_name"]
        course_name=request.form["course"]

        stud_class=Classes.query.all()

        for x in stud_class:
            if class_name==x.classname:
                flash(1)
                return redirect(url_for("class_reg"))


        course_name=Course.query.filter_by(course_name=course_name).first()
        classes=Classes(classname=class_name,camera_name=camera_name,course_class=course_name)
        db.session.add(classes)
        db.session.commit()
        flash("Class Added Successfully!!")
        return redirect(url_for('class_reg'))


###### Update course in the database
@app.route('/update_course',methods=["GET","POST"])
def update_courses():
    if request.method=="POST":
        update_query_course=Course.query.get(request.form.get('id'))
        update_query_course.course_name=request.form["course"]
        db.session.commit()
        flash("Course updated Sucessfully!!!")
        return redirect(url_for('course_reg'))

##############Update lec times
@app.route('/update_timing',methods=["GET","POST"])
def update_timings():
    if request.method=="POST":
        update_query_course=timing.query.get(request.form.get('id'))
        update_query_course.hours=request.form["hours"]
        update_query_course.minutes=request.form["minutes"]
        db.session.commit()
        flash("Timing updated Sucessfully!!!")
        return redirect(url_for('lecture_timings'))

###### Update class in the database
@app.route('/update_classes',methods=["GET","POST"])
def update_class():
    if request.method=="POST":
        update_query_class=Classes.query.get(request.form.get('id'))
        update_query_class.classname=request.form["class_name"]
        update_query_class.camera_name=request.form["camera_ip"]
        new_temp=request.form["course_name"]
        temp1=Course.query.filter_by(course_name=new_temp).first()
        update_query_class.course_sel=temp1.course_id

        db.session.commit()
        flash("Class updated Sucessfully!!!")
        return redirect(url_for('class_reg'))


###### Update student  in the database
@app.route('/update',methods=["GET","POST"])
def update():
    if request.method=="POST":
        update_query=Students.query.get(request.form.get('id'))
        #update_query_class=Students.query.get(request.form.get('classid'))
        update_query.rank=request.form['rank']
        update_query.first_name=request.form['first_name']
        update_query.last_name=request.form['last_name']
        update_query.phone=request.form['phone']
        temp=request.form["course"]
        # temp=request.form["class_name"]
        #print(temp)
        temp1=Course.query.filter_by(course_name=temp).first()
        update_query.course_sel=temp1.course_id

        db.session.commit()
        flash("Student updated Sucessfully!!")
        return redirect(url_for('index'))


@app.route('/update_attendance',methods=["GET","POST"])
def update_attendance():
    if request.method=="POST":
        update_query=arecord.query.get(request.form.get('id'))
        #print(update_query)
        #update_query.attend=request.form['value_attend']
        temp=request.form['value_attend']
        if temp=="True":
            update_query.attend=True
            #print("0")
        else:
            update_query.attend=False
        #print(bool(request.form['value_attend']))
        #update_query_course.course_name=request.form["course"]
        db.session.commit()
        flash("Attendace updated Sucessfully!!!")
        return redirect(url_for('attendance_records'))



# @app.route('/update_lecture',methods=["GET","POST"])
# def update_lec():
#     if request.method=="POST":
#         # update_query=Lectures.query.get(request.form.get('id'))
#         # update_query.lecture_name=request.form['name']
#         # update_query.lecture_day=request.form['day']
#         # update_query.lecture_time=request.form['time']
#         # temp=request.form["class_name"]
#         # lecture_id=request.form.get('id')
#         lecture_id=request.form.get('id')
#         # Lectures.query.filter_by(lecture_id=lecture_id).delete()
#         delete_student=Lectures.query.get(lecture_id)
#         db.session.delete(delete_student)
#         db.session.commit()
#         lecture_name=request.form["name"]
#         lecture_day=request.form["day"]
#         lecture_time=request.form["time"]
#         class_name=request.form["class_name"]
#         lec=Lectures(lecture_id=lecture_id,lecture_name=lecture_name,lecture_day=lecture_day,lecture_time=lecture_time)
#         db.session.add(lec)
#         db.session.commit()
#         class_name_obj=Classes.query.filter_by(classname=class_name).first()
#         lec.subjects.append(class_name_obj)
#         db.session.commit()

#         # temp1=Classes.query.filter_by(classname=temp).first()
        
#         # for i,j in enumerate(update_query.subjects):
            

        
#         flash("Lecture updated Sucessfully!!")
#         return redirect(url_for('lecture_reg'))


##### Delete student in the database
@app.route('/delete/<id>')
def delete(id):
    delete_student=Students.query.get(id)
    course=delete_student.course_sel

    course_1=Course.query.filter_by(course_id=course).first()
    a=course_1.course_name
    b=delete_student.roll_no
    # shutil.rmtree("static/photo/"+str(a)+"/"+str(b)+"jpg",ignore_errors = True)
    os.remove("static/photo/"+str(a)+"/"+str(b)+".jpg")
    if os.path.isfile('static/embeddings/'+str(a)+'.dat'):
        with open('static/embeddings/'+str(a)+'.dat',"rb") as f:
            encoded = pickle.load(f)
        with open('static/embeddings/'+str(a)+'.dat', 'wb') as f1:
            del encoded[str(a)+"_"+str(b)]
            
            pickle.dump(encoded,f1)

    db.session.delete(delete_student)
    db.session.commit()
    flash("Student Deleted Sucessfully!!")
    return redirect(url_for('index'))

##### Delete class in the database
@app.route('/delete_classes/<id>')
def delete_class(id):
    delete_class=Classes.query.get(id)
    db.session.delete(delete_class)
    db.session.commit()
    flash("Class Deleted Sucessfully!!")
    return redirect(url_for('class_reg'))

@app.route('/delete_timings/<id>')
def delete_timings(id):
    delete_tim=timing.query.get(id)
    db.session.delete(delete_tim)
    db.session.commit()
    flash("Timing Deleted Sucessfully!!")
    return redirect(url_for('lecture_timings'))

##### Delete course in the database
@app.route('/delete_course/<id>')
def delete_courses(id):
    delete_course=Course.query.get(id)
    os.remove('static/embeddings'+'/'+str(delete_course.course_name)+'.dat')
    location="static/photo/"
    path=os.path.join(location,str(delete_course.course_name))
    print(path)
    shutil.rmtree(path,ignore_errors = True)
    db.session.delete(delete_course)
    db.session.commit()
    flash("Course Deleted Sucessfully!!")
    return redirect(url_for('course_reg'))


# @app.route('/delete_lecture/<id>')
# def delete_lec(id):
#     delete_lecture=Lectures.query.get(id)
#     db.session.delete(delete_lecture)
#     db.session.commit()
#     flash("Lecture Deleted Sucessfully!!")
#     return redirect(url_for('lecture_reg'))


# @app.route('/lecture_registered/',methods=["GET","POST"])
# def lec_complete_reg():
#     if request.method=="POST":
#         lecture_name=request.form["name"]
#         lecture_day=request.form["day"]
#         lecture_time=request.form["time"]
#         class_name=request.form["class_name"]
#         lec=Lectures(lecture_name=lecture_name,lecture_day=lecture_day,lecture_time=lecture_time)
#         db.session.add(lec)
#         db.session.commit()
#         class_name_obj=Classes.query.filter_by(classname=class_name).first()
#         lec.subjects.append(class_name_obj)
#         db.session.commit()
#         flash("Lecture Added Sucessfully!!")

#         return redirect(url_for('lecture_reg'))
        

#### After entering the details of the student session created to pass the contents of students to the next pages
@app.route('/index_2/', methods =["GET", "POST"])
def indexing():
    if request.method=="POST":
        roll_no=request.form["roll"]
        rank=request.form["rank"]
        first_name=request.form["first_name"]
        last_name=request.form["last_name"]
        phone=request.form["phone"]
        course=request.form["course"]
        # stream=request.form["stream"]
        # class_name=request.form["class_name"]

        studs=Students.query.all()

        for x in studs:
            if roll_no==x.roll_no:
                flash(1)
                return redirect(url_for("index"))
        

        session["roll_no"]=roll_no
        session["rank"]=rank
        session["first_name"]=first_name
        session["last_name"]=last_name
        session["phone"]=phone
        session["course"]=course
        # session["stream"]=stream
        
        # session["class_name"]=class_name
        # 
        return render_template('index.html',roll_no=roll_no,rank=rank,first_name=first_name,last_name=last_name,phone=phone,course=course)


### Camera access
def gen(camera):
    while True:
        frame = camera.get_feed()
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')

            
@app.route('/video_feed/')
def video_feed():
    camera = get_camera()
    ip='rtsp://admin:admin@123@192.168.1.240:554/cam/realmonitor?channel=4&subtype=0'
    #t1 = threading.Thread(target=camera.start_cam, args=(ip,))
    #t1.start()
    camera.start_cam(ip)
    return Response(gen(camera),
        mimetype='multipart/x-mixed-replace; boundary=frame')


####Timestamp Creation
@app.route('/capture/')
def capture():
    #print(name)
    camera = get_camera()
    course=session.get("course")
    roll_no=session.get("roll_no")
    stamp,_ = camera.capture(course,roll_no)
    #print(filename)
    #f = ('%s.jpeg' % time.strftime("%Y%m%d-%H%M%S"))
    #camera.save('%s/%s' % ('None_None', f))

    return redirect(url_for('show_capture', timestamp=stamp))
    

"""    
@app.route('/uploads/<path:filename>', methods=['GET', 'POST'])


def download(filename):
    filename=str(request.args.get('first_name'))
    uploads = os.path.join(current_app.root_path, app.config['UPLOAD_FOLDER'])
    return send_from_directory(directory=uploads, filename=filename)
"""    
def stamp_file(timestamp):
    roll_no=session.get("roll_no")
    first_name=session.get("first_name")
    last_name=session.get("last_name")
    roll_no=session.get("roll_no")
    course=session.get("course")
    return 'photo/'+course+'/'+roll_no+'.jpg'


##### Complete Details of students before inserting to the Database     
@app.route('/capture/image/<timestamp>', methods=['POST', 'GET'])
def show_capture(timestamp):
    path = stamp_file(timestamp)
    # print(path)
    roll_no=session.get("roll_no")
    rank=session.get("rank")
    first_name=session.get("first_name")
    last_name=session.get("last_name")
    phone=session.get("phone")
    # stream=session.get("stream")
    course=session.get("course")
    return render_template('capture.html', path=path,roll_no=roll_no,rank=rank,first_name=first_name,last_name=last_name,phone=phone,course=course)

#######################get query and access attendance
"""def send_attendance():
    for i in os.listdir('/static/attendance'):
        if i.endswith(".xls"):"""
       


"""
@app.route('/capture/image/<timestamp>', methods=['POST', 'GET'])
def show_capture(timestamp):
    
    path = stamp_file(timestamp)


    #email_msg = None
    #if request.method == 'POST':
        

    return render_template('capture.html',
        stamp=timestamp, path=path)

</form>
 <form method="GET" action="{{url_for('index')}}">
<button type="submit" > Take photo </button>
</form>"""

###################################################
####added by s

@app.route('/test')
def test():
    return render_template('program.html')



@app.route('/foo', methods=['POST'])
def foo():
    flag=False
    global video
    global marked_courses
    
    f = datetime.now()
    global course_current
    #dont delete, parses through all the ips of camera
    course_name=[]
    class_index=[]
    if Classes.query.order_by(Classes.camera_name).all() is not None:
        if len(marked_courses)==len(Classes.query.order_by(Classes.camera_name).all()):
            marked_courses=[]
    
        for class_ip in Classes.query.order_by(Classes.camera_name).all():
            
            if class_ip.camera_name not in marked_courses:
               
                marked_courses.append(class_ip.camera_name)
                #print(marked_courses)

                class_index.append(class_ip.camera_name)
                if class_ip.camera_name=='0':
                    video=cv2.VideoCapture(int(class_ip.camera_name))
                else:

                    video=cv2.VideoCapture(class_ip.camera_name)
                course_current=Course.query.filter(Course.course_id==class_ip.course_sel)
                #print(class_ip.course_sel)
                for p in course_current:
                
                    s=p.course_name
                    course_name.append(s)
                    #print(course_name)
                #t1 = threading.Thread(target=program, args=(flag,s))
                #t1.start()
                
                program(flag,s)
            
            
        
    
    #video =cv2.VideoCapture(0)
    
    
    #program(flag,class_ip)
    
    return foo() 

@app.route('/new', methods=['POST'])
def new():
    video.release()
    # flag=True
    
    # program(flag,course_current)
    
    return redirect(url_for('attendance_records'))
@app.route('/highway', methods=['POST','GET'])
def highway():
    return redirect(url_for('index'))

###################program to convert data from sql to dict
#def check_unknown_id():
def to_dict(row):
    if row is None:
        return None

    rtn_dict = dict()
    keys = row.__table__.columns.keys()
    for key in keys:
        #print(key)
        rtn_dict[key] = getattr(row, key)
        #print(rtn_dict)
    return rtn_dict
##################program to get date input from user 
@app.route('/excel', methods =["GET", "POST"])
def excel():
    
    if request.method == "POST":
       # getting input with name = fname in HTML form
       start = request.form.get("startd")
       # getting input with name = lname in HTML form 
       end = request.form.get("endd") 
       select = request.form.get('coursed')
       number = request.form.get('lec_nod')
       
       #print(start)
       #print(end)
       qry = arecord.query.filter(and_(arecord.date.between(start, end),arecord.name_a.like(select),arecord.lecture_no.like(number))).all()
       #print(qry)
       if qry==[]:
           return redirect(url_for("attendance_records")) 
    
       data_list = [to_dict(item) for item in qry]
       df = pd.DataFrame(data_list)
       #print(df)
       #a=df.columns
       df.drop('name_a',inplace=True,axis=1)
       df.drop('primkey',inplace=True,axis=1)
       df.drop('lecture_no',inplace=True,axis=1)
       s = df.groupby(['id_a']).cumcount()

       df1 = df.set_index(['id_a', s]).unstack().sort_index(level=1, axis=1)
       df1.columns = [f'{x}{y}' for x, y in df1.columns]
       df1 = df1.reset_index()
       #print (df1)
       #print(df.date.unique())
       #print(df)
       resp = make_response(df1.to_csv())
       resp.headers["Content-Disposition"] = "attachment; filename=export.csv"
       resp.headers["Content-Type"] = "text/csv"
       return resp
       
       #for i in qry:
       #print(i.date)
    return redirect(url_for("attendance_records"))  
#return render_template("attend_page.html")
recognition_t=0.6
confidence_t=0.99


encoder_model = 'facenet_keras.h5'

#detector=MTCNN()
detector=MTCNN()
face_encoder = load_model(encoder_model)
directory='static/embeddings'
encoded={}
def send_encodings(directory):
    encoded={}
    for filename in os.listdir(directory):
        if filename.endswith(".dat"):
            if os.path.isfile('static/embeddings/'+str(filename)):
                with open('static/embeddings/'+str(filename),"rb") as f:
                    e= pickle.load(f)
                    encoded.update(e)
    return encoded
def get_encode(face_encoder, face, size):
    face = normalize(face)
    face = cv2.resize(face, size)
    encode = face_encoder.predict(np.expand_dims(face, axis=0))[0]
    return encode

def get_face(img, box):
    [[x1, y1, width, height]] = box
    x1, y1 ,x2,y2= int(x1), int(y1),int(width),int(height)
    #x2, y2 = x1 + width, y1 + height
    face = img[y1:y2, x1:x2]
    return face, (x1, y1), (x2, y2)
 
def normalize(img):
    mean, std = img.mean(), img.std()
    return (img - mean) / std

l2_normalizer = Normalizer('l2')



### collect daywise attendance by checking through a list of ppl

def attendance_in_db(a,t,lec_no,course_current):
    encoded=send_encodings(directory)
    # print("marking attendance")
    # print(a)
    #date_p=datetime.datetime.now()
    flag_a=1
    #print(course_current)
    if len(a)==0:
        for person_name in encoded:
            f=str(person_name).split('_')
            if str(person_name) not in a and f[0]==str(course_current):
                #print("no one is there")
                l=str(person_name).split('_')
                #print(l)
                flag_a=0
                marked =arecord(id_a=l[1],primkey=None,name_a=l[0],lecture_no=lec_no,attend=False) 
                db.session.add(marked)
                #classes=Classes(classname=class_name,camera_name=camera_name,course_class=course_name), date=date_p.strftime("%x")
                db.session.commit()


   

        
    if len(a)>0:
        #print("test")
        for person_name in encoded:
            #print(person_name)
            #print(a)

            #for x in range(0,len(a)):
                #spl=str(a[x]).split('_')
                #cou=spl[0]
                #print(spl[0])
            f=str(person_name).split('_')
            #print("course going on is "+ course_current)
            if str(person_name) in a and f[0]==str(course_current):
                flag_a=0
                
            
                
                marked =arecord(id_a=f[1],primkey=None, name_a=f[0],lecture_no=lec_no,attend=True) 
                #classes=Classes(classname=class_name,camera_name=camera_name,course_class=course_name)
                db.session.add(marked)
                db.session.commit()
                #print("attendance marked true!")
                #ResultProxy = db.session.execute(query)
                #query = db.insert(emp) 
            if str(person_name) not in a and f[0]==str(course_current):
                #print("else vala test")
                l=str(person_name).split('_')
                #print(l)
                flag_a=0
                marked =arecord(id_a=f[1],primkey=None, name_a=f[0],lecture_no=lec_no,attend=False) 
                db.session.add(marked)
                #classes=Classes(classname=class_name,camera_name=camera_name,course_class=course_name), date=date_p.strftime("%x")
                db.session.commit()
                #print(l)
                #query = db.insert(emp).values(Id=l[1], name=l[0], date=datetime(2015, 6, 5, 8, 10, 10, 10), attendance=False) 
                #ResultProxy = db.session.execute(query)
    return flag_a                
    #results = db.session.execute(db.select([emp])).fetchall()
    #df = pd.DataFrame(results)
    #df.columns = results[0].keys()
    #print(df.head(2))

    
def mark_attendance_of_a_lec(a,t,lec_no):
    #workbook = xlwt.Workbook() 
    course=check_which_course(a)
    
    #change this in pandas df to open and edit excel sheet 
    
    encoded=send_encodings(directory)

    workbook = xlwt.Workbook()  	 
    sheet = workbook.add_sheet(str(t.year)+"_"+str(t.month)+"_"+str(t.day)) 
    sheet.write(0,0,"Name")
    row = 1
    col = 0
    if len(a)>0:
        for person_name in encoded:
            #print(person_name)
            #print(a)

            for x in range(0,len(a)):
                spl=str(a[x]).split('_')
                cou=spl[0]
                if person_name in a:
                
                    l=str(a[x]).split('_')
                    #print(l)
                    sheet.write(row, col,     str(l[0]))
    
                    sheet.write(row, col+1,     str(l[1]))
                    sheet.write(row,col+2,"A")
                    
                else: 
                    l=str(person_name).split('_')
                    sheet.write(row, col,     str(l[0]))
    
                    sheet.write(row, col+1,     str(l[1]))
                    sheet.write(row,col+2,"A")
                row+=1
        if not os.path.exists('static/attendance/'+str(course)):
        
            os.makedirs('static/attendance/'+str(course))
        workbook.save('static/attendance/'+str(course)+"/"+str(t.day)+"_"+str(t.month)+"_"+str(t.year)+"_"+str(t.hour)+":"+str(t.minute)+".xls")
        
        
       
        
       
  
def check_which_course(a):
    number_of_s={}
    for x in range(0,len(a)):
        l=str(a[x]).split('_')
        if l[0] not in number_of_s:
            number_of_s[l[0]]=1
        else:
            number_of_s[l[0]]+=1
    course = max(number_of_s, key=number_of_s.get)
    return course
        




classNames = []
with open('coco.names','r') as f:
    classNames = f.read().splitlines()
#print(classNames)
thres = 0.5 # Threshold to detect object
nms_threshold = 0.2 #(0.1 to 1) 1 means no suppress , 0.1 means high suppress
weightsPath = "frozen_inference_graph.pb"
configPath = "ssd_mobilenet_v3_large_coco_2020_01_14.pbtxt"
net = cv2.dnn_DetectionModel(weightsPath,configPath)
net.setInputSize(320,320)
net.setInputScale(1.0/ 127.5)
net.setInputMean((127.5, 127.5, 127.5))
net.setInputSwapRB(True)

    

            
def program(flag,course_current):
    encoded=send_encodings(directory)
    present_candidates=[]
    flag_a=0
    
    while True:
        check,frame=video.read()
        #frame=cv2.resize(frame,(224,224))
        total_people=0
        t=datetime.now()
        #frame=sr.upsample(frame)
        #total_frames = total_frames + 1
        #print(t.second)

        faces,_=detector.detect(frame)
        classIds, confs, bbox = net.detect(frame,confThreshold=thres)
        bbox = list(bbox)
        confs = list(np.array(confs).reshape(1,-1)[0])
        confs = list(map(float,confs))
    
        indices = cv2.dnn.NMSBoxes(bbox,confs,thres,nms_threshold)
        if len(classIds) != 0:
        
            for i in indices:
                i = i[0]
            
                if classIds[i][0]==0:
                    total_people+=1
            
    
        #print(faces)
        if faces is not None:
            #print("face detected")
            for person in faces:
                bounding_box=person
                face, pt_1, pt_2 = get_face(frame, [bounding_box])
                encode = get_encode(face_encoder, face,(160,160))
                encode = l2_normalizer.transform(encode.reshape(1, -1))[0]
                name = 'unknown'
                distance = float("inf")
                for (db_name, db_enc) in encoded.items():
                    dist = cosine(db_enc, encode)
                    if dist < recognition_t and dist < distance:
                        name = db_name
                        distance = dist
                        if name not in present_candidates:
                            present_candidates.append(name)
                course="temp"
                if len(present_candidates)!=0:
                    course=check_which_course(present_candidates)
                    #print(course)
                """if name=='unknown' and t.second==24:
                    name_l=random.random()
                    present_candidates.append(name_l)
                    #classdetails=Classes.query.filter_by(camera_name=str(classroom_ip)).first()
                    #namef=classdetails.course_id

                    #save encodings in temp file
                    if os.path.exists('static/embeddings/'+"temp"+'.dat'):
            
                        with open('static/embeddings/'+"temp"+'.dat',"rb") as f:
                            randencode = pickle.load(f)
              



                        with open('static/embeddings/'+"temp"+'.dat', 'wb') as f1:
            
                            randencode[str(course)+"_"+str(name_l)]=encode
                            pickle.dump(randencode,f1)
                    else:
                        randencode={}
                        with open('static/embeddings/'+"temp"+'.dat', 'wb') as f2:
            
                            randencode[str(course)+"_"+str(name_l)]=encode
                            pickle.dump(randencode,f2)
                    if not os.path.exists('static/photo/temp'):
                        os.makedirs('static/photo/temp')
                    filename ='static/photo/temp'+'/'+ str(name_l)+".jpg"
                    encoded=send_encodings(directory)

        
                    if not cv2.imwrite(filename, frame):
                        raise RuntimeError("Unable to capture image "+timestamp)"""
        #total_cam=len(Classes.query.order_by(Classes.camera_name).all())
        #i+=1
        if t.minute==30:
            marked_courses=[]
            #i=0
        ######## code to test asap
        # if  t.second==2 or t.second==18:
        #     flag_a=1           
        # if  4<t.second<20 and flag_a==1 : 
            

           
        #     lec_no=t.minute
            
        #     flag_a=attendance_in_db(present_candidates,t,lec_no,course_current) 
        #     return foo() 
        # if timing.query.order_by(timing.time_id).all() is not None:
        #     for sch in timing.query.order_by(timing.time_id).all():
        #         a=sch.time.split(":")
        #         print(t.minute)
        #         print(a[1])
        #         if t.hour==int(a[0]) and t.minute==int(a[1]) and t.second==1:
        #             flag_a=1           
        #         if t.hour==int(a[0]) and t.minute==int(a[1]) and 30<t.second<59 : 
           
        #             lec_no=int(sch.time_id)
            
        #             flag_a=attendance_in_db(present_candidates,t,lec_no,course_current) 
        #             return foo()


                

        #############8:15-9 am 
        if t.hour==8 and t.minute==20 and t.second==22:
            flag_a=1           
        if t.hour==8 and t.minute==20 and t.second==25 and flag_a==1 : 
           
            lec_no=[True,0,0,0,0,0,0,0]
            
            flag_a=attendance_in_db(present_candidates,t,lec_no,course_current) 
            return
        #############9 - 9:45 am
        if t.hour==9 and t.minute==5 and t.second==22:
            flag_a=1           
        if t.hour==9 and t.minute==5 and t.second==25 and flag_a==1 : 
           
            lec_no=[True,0,0,0,0,0,0,0]
            
            flag_a=attendance_in_db(present_candidates,t,lec_no,course_current) 
            return
        ############# 9:50-10:30
        if t.hour==9 and t.minute==55 and t.second==22:
            flag_a=1           
        if t.hour==9 and t.minute==55 and t.second==25 and flag_a==1 : 
           
            lec_no=[True,0,0,0,0,0,0,0]
            
            flag_a=attendance_in_db(present_candidates,t,lec_no,course_current) 
            return
        #################10:30-11:30
        if t.hour==10 and t.minute==35 and t.second==22:
            flag_a=1           
        if t.hour==10 and t.minute==35 and t.second==25 and flag_a==1 : 
           
            lec_no=[True,0,0,0,0,0,0,0]
            
            flag_a=attendance_in_db(present_candidates,t,lec_no,course_current) 
            return
        ############11:40-12:25
        if t.hour==11 and t.minute==45 and t.second==22:
            flag_a=1           
        if t.hour==11 and t.minute==45 and t.second==25 and flag_a==1 : 
           
            lec_no=[True,0,0,0,0,0,0,0]
            
            flag_a=attendance_in_db(present_candidates,t,lec_no,course_current) 
            return
        #################12:30-1:10
        if t.hour==12 and t.minute==35 and t.second==22:
            flag_a=1           
        if t.hour==12 and t.minute==35 and t.second==25 and flag_a==1 : 
           
            lec_no=[True,0,0,0,0,0,0,0]
            
            flag_a=attendance_in_db(present_candidates,t,lec_no,course_current) 
            return
        ##########1:15-2
        if t.hour==13 and t.minute==20 and t.second==22:
            flag_a=1           
        if t.hour==13 and t.minute==20 and t.second==25 and flag_a==1 : 
           
            lec_no=[True,0,0,0,0,0,0,0]
            
            flag_a=attendance_in_db(present_candidates,t,lec_no,course_current) 
            return
        """if t.second==59:
            mark_attendance_of_a_lec(present_candidates,t) 
            course=check_which_course(present_candidates)
            
            if not os.path.exists('static/proof/'+str(course)):
                
                os.makedirs('static/proof/'+str(course))
            filename ='static/proof/'+str(course)+'/'+ str(t.day)+"_"+str(t.month)+"_"+str(t.year)+"_"+str(t.hour)+":"+str(t.minute)+".jpg"
            query

        
            if not cv2.imwrite(filename, frame):
                raise RuntimeError("Unable to capture image "+timestamp)   """ 
        if flag:
           break  
     
    video.release()



if __name__ == '__main__':
    

    app.run(host='0.0.0.0', port=5010, debug=True)
